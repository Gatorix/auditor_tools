import xlrd
import xlwt
import os
import msvcrt
import win32com.client as win32
import numpy as np
import pandas as pd
import glob
import openpyxl


class Merge():
    def __init__(self):
        self.rdfilepath = glob.glob(r"./*试算表*.xls*")
        self.head_line = [u'科目代码', u'科目名称', u'期初余额', u'本期借方', u'本期贷方', u'期末余额']
        # self.rd1filepath = glob.glob(r"./*科目余额表*.xls*")

    def exit_with_anykey(self):
        print("按任意键退出")
        ord(msvcrt.getch()) 
        os._exit(1)

    def formatXLS(self, filepath):  # 转换为xlsx格式
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(filepath)

        # FileFormat = 51 is for .xlsx extension
        # FileFormat = 56 is for .xls extension

        wb.SaveAs(filepath+"x", FileFormat=51)
        wb.Close()
        excel.Application.Quit()

    def get_col_value(self, sheet, col, start_row):  # 获取表格列值
        return sheet.col_values(col, start_row)

    def format_code_list(self, code):  # 转换科目代码类型
        int_code = []
        for i in range(len(code)):
            int_code.append(str(int(code[i])))
        return int_code

    def g_first_level_code(self, third_level_code):  # 获取一级科目代码
        first_level_code = []
        for i in range(len(third_level_code)):
            first_level_code.append(third_level_code[i][:4:])
        return first_level_code

    def g_second_level_code(self, third_level_code):  # 获取二级科目代码
        second_level_code = []
        for i in range(len(third_level_code)):
            second_level_code.append(third_level_code[i][:6:])
        return second_level_code

    def g_first_level_name(self, third_level_name):  # 获取一级科目名称
        first_level_name = []
        for i in range(len(third_level_name)):
            first_level_name.append(third_level_name[i].split("-")[0])
        return first_level_name

    def g_second_level_name(self, third_level_name):  # 获取二级科目名称
        second_level_name = []
        for i in range(len(third_level_name)):
            if len(third_level_name[i].split("-")) == 3:
                second_level_name.append(third_level_name[i].rsplit("-", 1)[0])
            elif len(third_level_name[i].rsplit("-", 1)) <= 2:
                second_level_name.append(third_level_name[i])
        return second_level_name

    def format_numbers(self, num_list):
        format_list = []
        for i in range(len(num_list)):
            try:
                format_list.append(float(num_list[i]))
            except ValueError:
                print("源表中'期初余额'、'借项'、'贷项'或'期末余额'字段中含有非数值，确认后重试")
                self.exit_with_anykey()
        return format_list

    def deduplication_level_code(self, level_code):
        n_level_code = list(set(level_code))
        n_level_code.sort(key=level_code.index)
        return n_level_code

    def replace_err_name(self, li, correct_name, error_name):
        for i in range(len(li)):
            for x in range(len(error_name)):
                if li[i] == error_name[x]:
                    li[i] = correct_name+li[i]
        return li

    def insert_err_name(self, li, correct_name, error_name):
        for i in range(len(li)):
            for x in range(len(error_name)):
                if li[i] == error_name[x]:
                    li[i] = li[i][:li[i].index('-'):] + \
                        correct_name+li[i][li[i].index('-')::]
        return li

    def write_list_nonformat(self, sheet, rows, cols, in_list):  # 写入list
        for i in range(rows):
            sheet.write(i+1, cols, in_list[i])

    def findStr(self, string, subStr, findCnt):
        listStr = string.split(subStr, findCnt)
        if len(listStr) <= findCnt:
            return -1
        return len(string)-len(listStr[-1])-len(subStr)

    def delete_err_name(self, li, correct_name, error_name):
        for i in range(len(li)):
            if li[i] == error_name:
                li[i] = li[i][:self.findStr(li[i], correct_name, 2):]
        return li

    def write_2d_list(self, sheet, t_Schema, start_row=0):  # 写入二维表
        style = xlwt.XFStyle()
        style.num_format_str = '#,##0.00'
        for row in range(len(t_Schema)):
            for col in range(0, len(t_Schema[row])):
                sheet.write(row+start_row, col, t_Schema[row][col], style)

    def write_header_line(self, head, sheet):
        for i in range(len(head)):
            sheet.write(0, i, head[i])

    def write_list(self, sheet, rows, cols, in_list):  # 写入list
        style = xlwt.XFStyle()
        style.num_format_str = '_ * #,##0.00_ ;_ * -#,##0.00_ ;_ * "-"??_ ;_ @_ '
        for i in range(rows):
            sheet.write(i+1, cols, in_list[i], style)

    def write_2d_list_opxl(self, data, sheet):
        for i in range(len(data)):
            for j in range(len(data[i])):
                sheet.cell(i+1, j+1, data[i][j])

    def main_function(self):
        try:
            print('检查文件路径……')
            print('打开试算表……')
            print(self.rdfilepath)
            self.workbook_r = xlrd.open_workbook(self.rdfilepath[0])  # 打开工作簿
            self.sheet_o_tb = self.workbook_r.sheet_by_index(0)  # 打开工作表
            print(self.sheet_o_tb)

        except FileNotFoundError:
            print('错误：未找到文件，确认后重试')
            self.exit_with_anykey()
        except PermissionError:
            print('错误：没有权限，关闭文件后重试')
            self.exit_with_anykey()
        # except IndexError:
        #     print('错误：首个工作表不是数据源')
        #     self.exit_with_anykey()

        print('读取科目名称及代码……')
        third_level_code = self.format_code_list(
            self.get_col_value(self.sheet_o_tb, 0, 1))  # 三级科目代码

        second_level_code = self.g_second_level_code(
            third_level_code)  # 二级科目代码

        first_level_code = self.g_first_level_code(third_level_code)  # 一级科目代码

        third_level_name = self.get_col_value(self.sheet_o_tb, 1, 1)  # 三级科目名称

        # 一级科目名称替换
        error_name_inventory = ['原材料', '材料采购', '原材料差异',
                                '半成品', '产成品', '发出商品-终端', '发出商品-解决方案', '发出商品-借货']
        correct_name_inventory = '存货-'

        error_name_captal = ['股本/实收资本', '库存股']
        correct_name_captal = '实收资本-'

        print('修改一级重码科目名称……')

        self.replace_err_name(
            third_level_name, correct_name_inventory, error_name_inventory)

        self.replace_err_name(
            third_level_name, correct_name_captal, error_name_captal)

        # 二级科目名称替换

        error_name_afterSale = ['工程施工-工程材料费',
                                '工程施工-工程分包费', '销售费用-工程材料费', '销售费用-工程分包费']
        correct_name_afterSale = '-售后服务'

        error_name_maintain = ['工程施工-工程杂费',
                               '工程施工-临时雇人人工费', '销售费用-工程杂费', '销售费用-临时雇人人工费']
        correct_name_maintain = '-维护费'

        error_name_trans = ['中转科目-待付员工款', '中转科目-跨供应商付款', '中转科目-跨客户收款',
                            '中转科目-零金额核销', '中转科目-银行购汇', '中转科目-应收退款', '中转科目-应收应付对冲','中转科目-期初导入']
        correct_name_trans = '-中转科目'

        error_name_others = ['工程施工-转销售费用', '工程施工-转营业成本', '销售费用-转受托开发支出']
        correct_name_others = '-其他'

        error_name_otherIncome = [
            '其他收益-软件退税', '其他收益-个税返还','其他收益-增值税返还', '其他收益-政府补助（本期收到）', '其他收益-政府补助（递延收益转入']
        correct_name_otherIncome = '-其他收益'

        error_name_tax = ['税金及附加-城市维护建设费', '税金及附加-地方教育费附加', '税金及附加-房产税',
                          '税金及附加-教育费附加', '税金及附加-其他', '税金及附加-土地使用税', '税金及附加-印花税']
        correct_name_tax = '-税金及附加'

        error_name_entrusted_rd = [
            '受托开发支出-研发费用转入', '受托开发支出-管理费用转入', '受托开发支出-销售费用转入', '受托开发支出-财务费用转入']
        correct_name_entrusted_rd = '-受托开发支出'

        error_name_rd = ['研发费用-转受托开发支出', '研发费用-转开发支出']
        correct_name_rd = '-转受托开发支出'

        error_name_dz = ['存货跌价准备-计提（原材料）', '存货跌价准备-计提（产成品）', '存货跌价准备-计提（半成品）']
        correct_name_dz = '-存货跌价准备'

        error_name_taxx = '应交税费-应交增值税-进项税额-不'
        correct_name_taxx = '-'

        print('修改二级重码科目名称……')

        self.delete_err_name(
            third_level_name, correct_name_taxx, error_name_taxx)

        self.insert_err_name(
            third_level_name, correct_name_dz, error_name_dz)

        self.insert_err_name(
            third_level_name, correct_name_afterSale, error_name_afterSale)

        self.insert_err_name(
            third_level_name, correct_name_maintain, error_name_maintain)

        self.insert_err_name(
            third_level_name, correct_name_trans, error_name_trans)

        self.insert_err_name(
            third_level_name, correct_name_others, error_name_others)

        self.insert_err_name(
            third_level_name, correct_name_otherIncome, error_name_otherIncome)

        self.insert_err_name(
            third_level_name, correct_name_tax, error_name_tax)

        self.insert_err_name(
            third_level_name, correct_name_rd, error_name_rd)

        self.insert_err_name(
            third_level_name, correct_name_entrusted_rd, error_name_entrusted_rd)

        second_level_name = self.g_second_level_name(
            third_level_name)  # 二级科目名称

        first_level_name = self.g_first_level_name(third_level_name)  # 一级科目名称
        print('读取试算表金额……')
        opening_balance = self.format_numbers(
            self.get_col_value(self.sheet_o_tb, 2, 1))  # 期初余额

        debit_amount = self.format_numbers(
            self.get_col_value(self.sheet_o_tb, 3, 1))  # 借方金额

        credit_amount = self.format_numbers(
            self.get_col_value(self.sheet_o_tb, 4, 1))  # 贷方金额

        ending_balance = self.format_numbers(
            self.get_col_value(self.sheet_o_tb, 5, 1))  # 期末余额

        # try:
        #     print('打开科目余额表……')
        #     self.workbook_r1 = xlrd.open_workbook(
        #         self.rd1filepath[0])  # 打开科目余额明细
        #     self.sheet_o_tb1 = self.workbook_r1.sheet_by_index(0)  # 打开工作表
        #     print(self.sheet_o_tb1)

        # except FileNotFoundError:
        #     print('错误：未找到文件，确认后重试')
        #     self.exit_with_anykey()
        # except PermissionError:
        #     print('错误：没有权限，关闭文件后重试')
        #     self.exit_with_anykey()
        # except IndexError:
        #     print('错误：首个工作表不是数据源')
        #     self.exit_with_anykey()
        # print('读取科目余额表……')

        # m_code = self.get_col_value(self.sheet_o_tb1, 2, 1)  # 明细科目代码

        # m_name = self.get_col_value(self.sheet_o_tb1, 5, 1)  # 明细科目名称

        # m_opening_balance = self.format_numbers(
        #     self.get_col_value(self.sheet_o_tb1, 15, 1))  # 明细期初余额

        # m_debit_amount = self.format_numbers(
        #     self.get_col_value(self.sheet_o_tb1, 16, 1))  # 明细借方金额

        # m_credit_amount = self.format_numbers(
        #     self.get_col_value(self.sheet_o_tb1, 17, 1))  # 明细贷方金额

        # m_ending_balance = self.format_numbers(
        #     self.get_col_value(self.sheet_o_tb1, 18, 1))  # 明细期末余额

        # print("创建明细科目原始DataFrame……")
        # o_detail = pd.DataFrame({
        #     'm_code': m_code,
        #     'm_name': m_name,
        #     'm_opening_balance': m_opening_balance,
        #     'm_debit_amount': m_debit_amount,
        #     'm_credit_amount': m_credit_amount,
        #     'm_ending_balance': m_ending_balance
        # })
        print('创建一级科目原始DataFrame……')
        # 原始一级科目
        o_first_level = pd.DataFrame({
            'first_level_code': first_level_code,
            'first_level_name': first_level_name,
            'opening_balance': opening_balance,
            'debit_amount': debit_amount,
            'credit_amount': credit_amount,
            'ending_balance': ending_balance
        })
        # 原始二级科目
        print('创建二级科目原始DataFrame……')
        o_second_level = pd.DataFrame({
            'second_level_code': second_level_code,
            'second_level_name': second_level_name,
            'opening_balance': opening_balance,
            'debit_amount': debit_amount,
            'credit_amount': credit_amount,
            'ending_balance': ending_balance
        })
        # 原始三级科目
        print('创建三级科目原始DataFrame……')
        o_third_level = pd.DataFrame({
            'third_level_code': third_level_code,
            'third_level_name': third_level_name,
            'opening_balance': opening_balance,
            'debit_amount': debit_amount,
            'credit_amount': credit_amount,
            'ending_balance': ending_balance
        })

        # 禁用pandas科学计数法
        pd.set_option('display.float_format', lambda x: '%.2f' % x)
        # 合并一级科目
        print('Group By一级科目……')
        first_level = o_first_level.groupby(by='first_level_code').sum()
        # 合并二级科目
        print('Group By二级科目……')
        second_level = o_second_level.groupby(by='second_level_code').sum()
        # 三级科目
        print('转换三级科目DataFrame至矩阵……')
        third_level_matrix = np.array(
            o_third_level).reshape(len(o_third_level), 6)

        # 明细科目
        # print('转换明细科目DataFrame至矩阵……')
        # m_detail_matrix = np.array(
        #     o_detail).reshape(len(o_detail), 6)
        
        first_level_value = np.array(first_level).tolist()

        second_level_value = np.array(second_level).tolist()

        first_level_mcode = np.array(self.deduplication_level_code(
            first_level_code)).reshape(len(first_level_value), 1)
        
        second_level_mcode = np.array(self.deduplication_level_code(
            second_level_code)).reshape(len(second_level_value), 1)

        first_level_mname = np.array(self.deduplication_level_code(
            first_level_name)).reshape(len(first_level_value), 1)
        
        second_level_mname = np.array(self.deduplication_level_code(
            second_level_name)).reshape(len(second_level_value), 1)
        
        if len(first_level_mcode) != len(first_level_mname):
            print('错误：存在一个科目代码对应多个科目名称的情况（一级科目），确认后重试')
            self.exit_with_anykey()
        elif len(second_level_mcode) != len(second_level_mname):
            print('错误：存在一个科目代码对应多个科目名称的情况（二级科目），确认后重试')
            self.exit_with_anykey()
        print('合并一级科目余额……')
        first_level_matrix = np.hstack(
            (first_level_mcode, first_level_mname, first_level_value))  # 合并一级科目余额表
        print('合并二级科目余额……')
        second_level_matrix = np.hstack(
            (second_level_mcode, second_level_mname, second_level_value))  # 合并二级科目余额表
        print('合并所有级次……')
        merge_matrix = np.vstack((
            first_level_matrix, second_level_matrix, third_level_matrix))

        merge_matrix_sort = merge_matrix[merge_matrix[:, 0].argsort()]  # 矩阵排序

        code = np.array([x[0] for x in merge_matrix_sort], dtype=str).tolist()

        name = np.array([x[1] for x in merge_matrix_sort], dtype=str).tolist()

        op_amount = np.array(
            [x[2] for x in merge_matrix_sort], dtype=float).tolist()

        de_amount = np.array(
            [x[3] for x in merge_matrix_sort], dtype=float).tolist()

        cr_amount = np.array(
            [x[4] for x in merge_matrix_sort], dtype=float).tolist()

        en_amount = np.array(
            [x[5] for x in merge_matrix_sort], dtype=float).tolist()

        merge_list = [code, name, op_amount, de_amount, cr_amount, en_amount]

        # receivables_index = [i for i, x in enumerate(
        #     merge_list[0]) if '1122' == merge_list[0][i][:4:]]#应收下标
        # prepaid_index = [i for i, x in enumerate(
        #     merge_list[0]) if '1123' == merge_list[0][i][:4:]]#预付下标
        # other_receivables=[i for i, x in enumerate(
        #     merge_list[0]) if '1221' == merge_list[0][i][:4:]]#其他应收下标

        # needed_subject=receivables_index+prepaid_index+other_receivables
        # n_code=[]
        # n_name=[]
        # n_op_amount=[]
        # n_de_amount=[]
        # n_cr_amount=[]
        # n_en_amount=[]
        # for i in range(len(needed_subject)):
        #     n_code.append(code[needed_subject[i]])
        #     n_name.append(name[needed_subject[i]])
        #     n_op_amount.append(op_amount[needed_subject[i]])
        #     n_de_amount.append(de_amount[needed_subject[i]])
        #     n_cr_amount.append(cr_amount[needed_subject[i]])
        #     n_en_amount.append(en_amount[needed_subject[i]])

        # n_merge_list=[n_code,n_name,n_op_amount,n_de_amount,n_cr_amount,n_en_amount]

        print('创建输出文件……')
        workbook_w = xlwt.Workbook()  # 创建文件

        sheet_w = workbook_w.add_sheet(
            '科目余额表', cell_overwrite_ok=True)
        print('写入Excel……')
        self.write_header_line(self.head_line, sheet_w)
        for i in range(6):
            self.write_list(sheet_w, len(code), i, merge_list[i])

        try:
            print('保存文件……')
            workbook_w.save('分级试算表.xls')  # 保存文件
            print('完成')
            self.exit_with_anykey()
        except PermissionError:
            print('错误：文件保存失败,关闭输出文件后重试')
            self.exit_with_anykey()

        # workbook = openpyxl.Workbook()

        # workbook_t = workbook.active
        # self.write_2d_list_opxl(merge_list,workbook_t)

        # workbook.save('test_3.xlsx')


if __name__ == "__main__":
    merge = Merge()
    merge.main_function()
